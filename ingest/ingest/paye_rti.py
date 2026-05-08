"""
PAYE RTI ingest — ONS non-seasonally-adjusted LA-level payroll data.

Fetches the current quarterly LAU file from ONS (discovering its URL from the
main RTI workbook), extracts October 2025 data for Kent + Medway, and writes:
  data/paye_rti/kent_medway_2025_10.parquet
  data/paye_rti/kent_medway_2025_10_meta.json
"""

import io
import json
import logging
from datetime import datetime, timezone
from pathlib import Path

import httpx
import openpyxl
import pandas as pd

log = logging.getLogger(__name__)

DATASET_PAGE = (
    "https://www.ons.gov.uk/employmentandlabourmarket/peopleinwork/"
    "earningsandworkinghours/datasets/"
    "realtimeinformationstatisticsreferencetablenonseasonallyadjusted"
)
MAIN_FILE_URL = (
    "https://www.ons.gov.uk/file?uri=/employmentandlabourmarket/peopleinwork/"
    "earningsandworkinghours/datasets/"
    "realtimeinformationstatisticsreferencetablenonseasonallyadjusted/"
    "current/rtinsaapr2026.xlsx"
)

KENT_MEDWAY_CODES = frozenset({
    "E06000035",  # Medway
    "E07000105",  # Ashford
    "E07000106",  # Canterbury
    "E07000107",  # Dartford
    "E07000108",  # Dover
    "E07000109",  # Gravesham
    "E07000110",  # Maidstone
    "E07000111",  # Sevenoaks
    "E07000112",  # Folkestone and Hythe
    "E07000113",  # Swale
    "E07000114",  # Thanet
    "E07000115",  # Tonbridge and Malling
    "E07000116",  # Tunbridge Wells
})

TARGET_PERIOD = "October 2025"
PERIOD_START = "2025-10-01"
PERIOD_END = "2025-10-31"
SOURCE_ID = "ons_paye_rti_nsa_la"

UA = (
    "WhereIsTheMoney/0.1 "
    "(github.com/CPritch/wheres-the-money; charlespritchard.work@gmail.com)"
)

DATA_DIR = Path(__file__).parents[2] / "data" / "paye_rti"


def _fetch(client: httpx.Client, url: str) -> bytes:
    log.info("GET %s", url)
    resp = client.get(url, follow_redirects=True)
    resp.raise_for_status()
    log.info("  %d bytes received", len(resp.content))
    return resp.content


def _find_lau_url(main_xlsx_bytes: bytes) -> str:
    """Extract the quarterly LAU data URL from the main RTI workbook."""
    wb = openpyxl.load_workbook(io.BytesIO(main_xlsx_bytes), read_only=False)
    ws = wb["19. Employees (LA)"]
    for row in ws.iter_rows(max_row=10):
        for cell in row:
            if cell.hyperlink and "rtinsa" in cell.hyperlink.target:
                return cell.hyperlink.target
    raise RuntimeError("LAU data URL not found in main RTI workbook")


def _extract_la_col(ws, target_period: str, kent_codes: frozenset) -> dict[str, dict]:
    """
    Read one LA-level sheet (employees or pay) and return a dict
    { lad_code: {"name": str, "value": int|float} } for the target period.
    """
    rows = list(ws.iter_rows(values_only=True))
    codes_row = rows[6]   # row 7: LAD codes
    names_row = rows[7]   # row 8: LAD names
    out: dict[str, dict] = {}
    for row in rows[8:]:
        if row[0] == target_period:
            for col_idx, code in enumerate(codes_row):
                if code in kent_codes and row[col_idx] is not None:
                    out[str(code)] = {
                        "name": names_row[col_idx],
                        "value": row[col_idx],
                    }
            break
    return out


def _parse_lau(lau_xlsx_bytes: bytes, fetched_at: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(io.BytesIO(lau_xlsx_bytes), read_only=True)
    employees = _extract_la_col(wb["19. Employees (LA)"], TARGET_PERIOD, KENT_MEDWAY_CODES)
    median_pay = _extract_la_col(wb["20. Median pay (LA)"], TARGET_PERIOD, KENT_MEDWAY_CODES)

    records = []
    for code in KENT_MEDWAY_CODES:
        emp = employees.get(code)
        pay = median_pay.get(code)
        if emp and pay:
            records.append({
                "lad_code": code,
                "lad_name": emp["name"],
                "employee_count": int(emp["value"]),
                "median_pay_gbp": float(pay["value"]),
                "total_payroll_estimate_gbp": float(emp["value"] * pay["value"]),
                "period_start": PERIOD_START,
                "period_end": PERIOD_END,
                "confidence": "measured",
                "source_id": SOURCE_ID,
                "fetched_at": fetched_at,
            })

    if len(records) != len(KENT_MEDWAY_CODES):
        missing = KENT_MEDWAY_CODES - {r["lad_code"] for r in records}
        log.warning("Missing data for %d LADs: %s", len(missing), missing)

    return pd.DataFrame(records)


def run(out_dir: Path = DATA_DIR) -> Path:
    parquet_path = out_dir / "kent_medway_2025_10.parquet"

    # If already fetched, skip re-download (data is historical — it won't change)
    if parquet_path.exists():
        log.info("Parquet already exists at %s — skipping fetch", parquet_path)
        return parquet_path

    out_dir.mkdir(parents=True, exist_ok=True)
    fetched_at = datetime.now(timezone.utc).isoformat()

    with httpx.Client(headers={"User-Agent": UA}, timeout=120) as client:
        main_bytes = _fetch(client, MAIN_FILE_URL)
        lau_url = _find_lau_url(main_bytes)
        log.info("LAU file URL: %s", lau_url)
        lau_bytes = _fetch(client, lau_url)

    df = _parse_lau(lau_bytes, fetched_at)

    meta = {
        "source_id": SOURCE_ID,
        "name": "Earnings and employment from PAYE RTI (non-seasonally adjusted, LA-level)",
        "publisher": "Office for National Statistics",
        "dataset_url": DATASET_PAGE,
        "file_url": lau_url,
        "fetched_at": fetched_at,
        "license": "OGL3",
        "methodology": (
            "Total payroll is approximated as employee count × median monthly pay "
            "for each local authority district. Using the median rather than the mean "
            "understates the true total for positively-skewed pay distributions. "
            "ASHE pay distributions will be used to refine this estimate in Phase 2."
        ),
    }
    meta_path = out_dir / "kent_medway_2025_10_meta.json"
    meta_path.write_text(json.dumps(meta, indent=2))

    df.to_parquet(parquet_path, index=False)
    log.info("Wrote %d rows → %s", len(df), parquet_path)
    return parquet_path
